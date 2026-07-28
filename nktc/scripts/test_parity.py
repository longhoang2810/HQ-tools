#!/usr/bin/env python3
import json
import shutil
import subprocess
import sys
import tempfile
import textwrap
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "nktc_process.py"
TEMPLATE = ROOT / "NKTC-xu-ly-excel.template.html"
EXCELJS = ROOT / "assets" / "exceljs.min.js"
NODE = shutil.which("node")

sys.path.insert(0, str(ROOT / "scripts"))
from nktc_process import to_float


HEADERS = [
    "So_to_khai",
    "Ma_LH",
    "Ma_DN_XNK",
    "Ten_DN_XNK",
    "Ma_dia_chi_DN_XNK",
    "So_quan_ly_cua_noi_bo_doanh_nghiep",
    "Tong_tri_gia_tinh_thue",
]

AMOUNT_CASES = [
    ("1,234.56", 1234.56),
    ("12.345,67", 12345.67),
    ("1.234.567,89", 1234567.89),
    ("1,234", 1234.0),
    ("1.234", 1.234),
    ("  9,999,999.99  ", 9999999.99),
    (1234.5, 1234.5),
    (0, 0.0),
    ("", 0.0),
    (None, 0.0),
    ("   ", 0.0),
    ("12 345", 0.0),
    ("#N/A", 0.0),
    ("abc", 0.0),
    ("1234 USD", 0.0),
]

UNPARSEABLE_VALUES = ["12 345", "#N/A", "abc", "1234 USD"]

STRICT_GRAMMAR_REJECTED = ["0x10", "1_000", "Infinity", "NaN", "abc", "12 345"]


NODE_HARNESS = r"""
const fs = require('fs');
const vm = require('vm');
const request = JSON.parse(fs.readFileSync(0, 'utf8'));
global.ExcelJS = require(request.exceljs);

const elements = new Map();
function element(id) {
  if (!elements.has(id)) {
    elements.set(id, {
      id,
      value: '',
      files: [],
      className: '',
      textContent: '',
      disabled: false,
      addEventListener() {},
      click() {},
    });
  }
  return elements.get(id);
}
global.document = {
  getElementById: element,
  createElement() { return {href: '', download: '', click() {}}; },
};
global.window = global;
global.confirm = () => true;
global.URL = {createObjectURL: () => 'blob:test', revokeObjectURL() {}};
global.setTimeout = fn => { fn(); return 0; };

element('regions').value = request.regions || 'Test\tHa Noi';
element('month').value = '06';
element('year').value = '2026';
element('tbMonth').value = '07';
element('tbYear').value = '2026';
element('outName').value = 'out.xlsx';
if (request.source) {
  element('source').files = [{
    arrayBuffer: async () => fs.promises.readFile(request.source),
  }];
}

const template = fs.readFileSync(request.template, 'utf8');
const match = template.match(/<script>\s*(\(\(\) => \{[\s\S]*?\}\)\(\);)\s*<\/script>\s*<\/body>/);
if (!match) throw new Error('Could not extract inline NKTC application JS');
const app = match[1].replace(
  /\}\)\(\);\s*$/,
  "globalThis.__NKTC_TEST__={num,extractRows,chooseRegion,buildRegionSheet,run,getUnparseableAmounts:()=>unparseableAmounts};})();"
);
vm.runInThisContext(app, {filename: request.template});
const api = global.__NKTC_TEST__;

(async () => {
  if (request.action === 'amounts') {
    process.stdout.write(JSON.stringify(request.values.map(value => api.num(value))));
    return;
  }
  if (request.action === 'amounts_with_count') {
    const values = request.values.map(value => api.num(value));
    process.stdout.write(JSON.stringify({values, unparseableAmounts: api.getUnparseableAmounts()}));
    return;
  }
  if (request.action === 'status') {
    await api.run();
    const status = element('status');
    process.stdout.write(JSON.stringify({kind: status.className, text: status.textContent}));
    return;
  }
  if (request.action === 'build') {
    const source = new ExcelJS.Workbook();
    await source.xlsx.load(fs.readFileSync(request.source));
    const rows = api.extractRows(source.worksheets[0]);
    const region = {name: 'Test', terms: ['ha noi']};
    const records = rows.filter(row => api.chooseRegion(row.A, [region]) === region.name);
    const output = new ExcelJS.Workbook();
    const count = api.buildRegionSheet(output, region.name, records);
    const data = await output.xlsx.writeBuffer();
    fs.writeFileSync(request.output, Buffer.from(data));
    process.stdout.write(JSON.stringify(count));
    return;
  }
  throw new Error(`Unknown action: ${request.action}`);
})().catch(error => {
  console.error(error && error.stack ? error.stack : error);
  process.exitCode = 1;
});
"""


def source_row(name, code, amount=100, declaration=100000001, address="Ha Noi"):
    return [
        declaration,
        "E21",
        code,
        name,
        address,
        f"12345678XK{declaration}",
        amount,
    ]


def write_source(path, rows, code_format=None):
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    if code_format:
        ws["C2"].number_format = code_format
    wb.save(path)


def inject_cached_formula_results(path, cached_results):
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", namespace)
    patched = path.with_name(f"{path.stem}-patched.xlsx")
    with zipfile.ZipFile(path, "r") as source_zip, zipfile.ZipFile(patched, "w") as target_zip:
        for info in source_zip.infolist():
            data = source_zip.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                root = ET.fromstring(data)
                for coordinate, (value, value_type) in cached_results.items():
                    cell = root.find(f".//{{{namespace}}}c[@r='{coordinate}']")
                    if cell is None:
                        raise AssertionError(f"Missing formula cell {coordinate}")
                    if value_type:
                        cell.set("t", value_type)
                    else:
                        cell.attrib.pop("t", None)
                    cached = cell.find(f"{{{namespace}}}v")
                    if cached is None:
                        cached = ET.SubElement(cell, f"{{{namespace}}}v")
                    cached.text = str(value)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
            target_zip.writestr(info, data)
    shutil.move(patched, path)


def write_regions(path):
    path.write_text("Test\tHa Noi\n", encoding="utf-8")


def run_python(source, output, regions):
    return subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            str(source),
            "-o",
            str(output),
            "--regions",
            str(regions),
            "--month",
            "06",
            "--year",
            "2026",
        ],
        check=True,
        capture_output=True,
        text=True,
    )


def region_snapshot(path):
    ws = load_workbook(path, data_only=True)["Test"]
    rows = []
    for row in range(5, ws.max_row + 1):
        rows.append(tuple(ws.cell(row, col).value for col in range(1, 7)))
    return rows, {str(rng) for rng in ws.merged_cells.ranges}


@unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
def run_js(temp_dir, request):
    harness = Path(temp_dir) / "nktc_js_harness.js"
    harness.write_text(textwrap.dedent(NODE_HARNESS), encoding="utf-8")
    payload = {
        "template": str(TEMPLATE),
        "exceljs": str(EXCELJS),
        **request,
    }
    result = subprocess.run(
        [NODE, str(harness)],
        input=json.dumps(payload, ensure_ascii=False),
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


class AmountParsingParityTest(unittest.TestCase):
    def test_python_amount_behaviour_table(self):
        for value, expected in AMOUNT_CASES:
            with self.subTest(value=value):
                self.assertEqual(to_float(value), expected)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_javascript_amount_behaviour_table(self):
        with tempfile.TemporaryDirectory() as tmp:
            actual = run_js(tmp, {"action": "amounts", "values": [v for v, _ in AMOUNT_CASES]})
        self.assertEqual(actual, [expected for _, expected in AMOUNT_CASES])

    def test_python_warning_counts_only_nonblank_unparseable_amounts(self):
        values = [0, "", None, "   ", *UNPARSEABLE_VALUES]
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            output = tmp_path / "python.xlsx"
            regions = tmp_path / "regions.txt"
            write_source(
                source,
                [source_row(f"Company {i}", str(i), value, 100000000 + i)
                 for i, value in enumerate(values, start=1)],
            )
            write_regions(regions)
            result = run_python(source, output, regions)
        self.assertIn(
            "CẢNH BÁO: 4 dòng có Trị giá không đọc được, đã ghi 0.00 - kiểm tra lại file nguồn.",
            result.stdout,
        )

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_javascript_warning_counts_only_nonblank_unparseable_amounts(self):
        values = [0, "", None, "   ", *UNPARSEABLE_VALUES]
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            write_source(
                source,
                [source_row(f"Company {i}", str(i), value, 100000000 + i)
                 for i, value in enumerate(values, start=1)],
            )
            status = run_js(
                tmp,
                {
                    "action": "status",
                    "source": str(source),
                    "regions": "Test\tHa Noi",
                },
            )
        self.assertEqual(status["kind"], "error")
        self.assertIn("4 dòng", status["text"])

    def test_python_strict_amount_grammar_and_count(self):
        stats = {}
        actual = [to_float(value, stats) for value in STRICT_GRAMMAR_REJECTED]
        self.assertEqual(actual, [0.0] * len(STRICT_GRAMMAR_REJECTED))
        self.assertEqual(stats.get("unparseable_amounts"), len(STRICT_GRAMMAR_REJECTED))
        for value, expected in AMOUNT_CASES[:7]:
            with self.subTest(value=value):
                self.assertEqual(to_float(value), expected)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_javascript_strict_amount_grammar_and_count(self):
        with tempfile.TemporaryDirectory() as tmp:
            actual = run_js(
                tmp,
                {"action": "amounts_with_count", "values": STRICT_GRAMMAR_REJECTED},
            )
            accepted = run_js(
                tmp,
                {"action": "amounts", "values": [value for value, _ in AMOUNT_CASES[:7]]},
            )
        self.assertEqual(actual["values"], [0.0] * len(STRICT_GRAMMAR_REJECTED))
        self.assertEqual(actual["unparseableAmounts"], len(STRICT_GRAMMAR_REJECTED))
        self.assertEqual(accepted, [expected for _, expected in AMOUNT_CASES[:7]])

    def test_python_boolean_amount_is_unparseable_and_counted(self):
        stats = {}
        self.assertEqual(to_float(True, stats), 0.0)
        self.assertEqual(stats.get("unparseable_amounts"), 1)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_javascript_boolean_amount_is_unparseable_and_counted(self):
        with tempfile.TemporaryDirectory() as tmp:
            actual = run_js(
                tmp,
                {"action": "amounts_with_count", "values": [True]},
            )
        self.assertEqual(actual, {"values": [0.0], "unparseableAmounts": 1})


class WorkbookParityTest(unittest.TestCase):
    def run_both_from_source(self, tmp_path, source):
        python_output = tmp_path / "python.xlsx"
        js_output = tmp_path / "javascript.xlsx"
        regions = tmp_path / "regions.txt"
        write_regions(regions)
        run_python(source, python_output, regions)
        js_count = run_js(
            tmp_path,
            {
                "action": "build",
                "source": str(source),
                "output": str(js_output),
            },
        )
        return python_output, js_output, js_count

    def run_both(self, tmp_path, rows, code_format=None):
        source = tmp_path / "source.xlsx"
        write_source(source, rows, code_format=code_format)
        return self.run_both_from_source(tmp_path, source)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_zero_padded_company_code_is_preserved_in_both(self):
        with tempfile.TemporaryDirectory() as tmp:
            python_output, js_output, _ = self.run_both(
                Path(tmp),
                [source_row("Company", 123456)],
                code_format="0000000000",
            )
            for output in (python_output, js_output):
                with self.subTest(output=output.name):
                    rows, _ = region_snapshot(output)
                    self.assertEqual(rows[0][2], "0000123456")

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_formula_codes_and_amounts_match_without_merging_companies(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            write_source(
                source,
                [
                    source_row("Company A", "=17+17", "=2+2", 100000001),
                    source_row("Company B", "=6+6", "=1+2", 100000002),
                ],
            )
            inject_cached_formula_results(
                source,
                {
                    "C2": (34, None),
                    "G2": (4, None),
                    "C3": (12, None),
                    "G3": (3, None),
                },
            )
            python_output, js_output, js_count = self.run_both_from_source(tmp_path, source)
            python_rows, _ = region_snapshot(python_output)
            js_rows, _ = region_snapshot(js_output)
        self.assertEqual(js_count["companies"], 2)
        self.assertEqual([row[2] for row in python_rows], ["34", "12"])
        self.assertEqual([row[5] for row in python_rows], [4, 3])
        self.assertEqual(js_rows, python_rows)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_formula_error_and_plain_error_cells_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
            ws.append(source_row("Company A", "=NA()", declaration=100000001))
            ws.append(source_row("Company B", "#DIV/0!", declaration=100000002))
            ws["C3"].data_type = "e"
            wb.save(source)
            inject_cached_formula_results(source, {"C2": ("#N/A", "e")})
            python_output, js_output, js_count = self.run_both_from_source(tmp_path, source)
            python_rows, _ = region_snapshot(python_output)
            js_rows, _ = region_snapshot(js_output)
        self.assertEqual(js_count["companies"], 2)
        self.assertEqual([row[2] for row in python_rows], ["#N/A", "#DIV/0!"])
        self.assertEqual(js_rows, python_rows)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_negative_zero_padded_company_code_matches_python_zfill(self):
        with tempfile.TemporaryDirectory() as tmp:
            python_output, js_output, _ = self.run_both(
                Path(tmp),
                [source_row("Company", -12)],
                code_format="0000",
            )
            for output in (python_output, js_output):
                with self.subTest(output=output.name):
                    rows, _ = region_snapshot(output)
                    self.assertEqual(rows[0][2], "-012")

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_same_code_nonadjacent_names_form_one_merged_company(self):
        rows = [
            source_row("CONG TY A", "0001", declaration=100000001),
            source_row("Cong ty A", "0001", declaration=100000002),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            python_output, js_output, js_count = self.run_both(Path(tmp), rows)
            self.assertEqual(js_count["companies"], 1)
            for output in (python_output, js_output):
                with self.subTest(output=output.name):
                    snapshot, merges = region_snapshot(output)
                    self.assertEqual([row[0] for row in snapshot], [1, None])
                    self.assertEqual(snapshot[0][1], "CONG TY A")
                    self.assertEqual(snapshot[0][2], "0001")
                    self.assertTrue({"A5:A6", "B5:B6", "C5:C6"}.issubset(merges))
            summary = load_workbook(python_output, data_only=True)["summary"]
            self.assertEqual(summary["D8"].value, 1)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_same_code_rows_are_bucketed_when_not_adjacent_after_sort(self):
        rows = [
            source_row("Zulu Company", "0001", declaration=100000001),
            source_row("Middle Company", "0002", declaration=100000002),
            source_row("Alpha Company", "0001", declaration=100000003),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            python_output, js_output, js_count = self.run_both(Path(tmp), rows)
            self.assertEqual(js_count["companies"], 2)
            for output in (python_output, js_output):
                with self.subTest(output=output.name):
                    snapshot, merges = region_snapshot(output)
                    self.assertEqual([row[0] for row in snapshot], [1, None, 2])
                    self.assertEqual([row[1] for row in snapshot], ["Alpha Company", None, "Middle Company"])
                    self.assertTrue({"A5:A6", "B5:B6", "C5:C6"}.issubset(merges))

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_blank_codes_are_separate_unmerged_companies(self):
        rows = [
            source_row("Company A", "", declaration=100000001),
            source_row("Company B", "   ", declaration=100000002),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            python_output, js_output, js_count = self.run_both(Path(tmp), rows)
            self.assertEqual(js_count["companies"], 2)
            for output in (python_output, js_output):
                with self.subTest(output=output.name):
                    snapshot, merges = region_snapshot(output)
                    self.assertEqual([row[0] for row in snapshot], [1, 2])
                    self.assertNotIn("A5:A6", merges)
                    self.assertNotIn("B5:B6", merges)
                    self.assertNotIn("C5:C6", merges)

    @unittest.skipUnless(NODE, "node is required for browser-JS parity tests")
    def test_vietnamese_name_order_and_stt_match_python(self):
        names = [
            "Âu Lạc",
            "Zeta",
            "ÁNH DƯƠNG",
            "Đại Phát",
            "an Binh",
            "Ô Môn",
            "bao Long",
        ]
        rows = [
            source_row(name, f"{index:04d}", declaration=100000000 + index)
            for index, name in enumerate(names, start=1)
        ]
        with tempfile.TemporaryDirectory() as tmp:
            python_output, js_output, _ = self.run_both(Path(tmp), rows)
            python_rows, _ = region_snapshot(python_output)
            js_rows, _ = region_snapshot(js_output)
        expected_names = sorted(names, key=str.lower)
        self.assertEqual([row[1] for row in python_rows], expected_names)
        self.assertEqual(
            [(row[0], row[1]) for row in js_rows],
            [(row[0], row[1]) for row in python_rows],
        )


if __name__ == "__main__":
    result = unittest.TextTestRunner(verbosity=2).run(
        unittest.defaultTestLoader.loadTestsFromModule(sys.modules[__name__])
    )
    if result.wasSuccessful():
        print("PASS_MARKER: parity")
    raise SystemExit(0 if result.wasSuccessful() else 1)
