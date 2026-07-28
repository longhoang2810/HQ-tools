#!/usr/bin/env python3
"""
NKTC Excel processor.

Step 1 - Filter & normalize a customs-declaration export:
  * keep rows where column B (Ma_LH) is in --ma-lh (default E21,G13)
  * sort by column I (Ten_DN_XNK) A->Z
  * remap columns:
        A = J  (Ma_dia_chi_DN_XNK)
        B = I  (Ten_DN_XNK)
        C = H  (Ma_DN_XNK)
        D = A  (So_to_khai)
        E = K  (So_quan_ly_cua_noi_bo_doanh_nghiep) with first 8 chars stripped
        F = P  (Tong_tri_gia_tinh_thue)

Step 2 - Build one workbook with one sheet per region plus unmatched (10 sheets total) by
  default. Use --separate-files to build one .xlsx per region plus unmatched instead. For each
  region, keep Step-1 rows whose column A (address) matches that region's
  terms. Matching is accent-insensitive: both accented ("Hà Nội") and
  unaccented ("Ha Noi") forms match, and "đ" is treated as "d". Each sheet/file
  is grouped by company (B+C) with merged STT/name/code cells, formatted header,
  borders and Times New Roman throughout. Regions (sheet/file <- address
  contains):
        hp <- hai ph, hai phong
        Hn <- ha noi
        PT <- phu tho, vinh phuc
        HY <- hung yen
        BN <- bac ninh, bac giang
        TH <- thanh hoa
        TQ <- tuyen quang
        QT <- quang tri
        NB <- nam dinh, ninh binh

Usage:
  python3 nktc_process.py INPUT.xlsx [-o OUTPUT.xlsx] [--sheet NAME]
         [--ma-lh E21,G13] [--header-rows 1] [--step1-out step1.xlsx]
         [--month MM] [--year YYYY]
         [--tb-no NUM] [--tb-day DD] [--tb-month MM] [--tb-year YYYY]

  # Legacy mode: write 9 separate .xlsx files into OUTDIR
  python3 nktc_process.py INPUT.xlsx -o OUTDIR --separate-files
"""
import argparse
import datetime
import os
import re
import sys
import unicodedata

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

# Source column indices (1-based)
SRC = {
    "So_to_khai":   1,   # A
    "Ma_LH":        2,   # B
    # Defaults for older exports; detect_src() below overrides these when
    # the workbook has recognizable headers.
    "Ma_DN_XNK":    8,   # H
    "Ten_DN_XNK":   9,   # I
    "Ma_dia_chi":  10,   # J
    "So_quan_ly":  11,   # K
    "Tong_tri_gia":16,   # P
    "Exclude_L":    12,   # L - skip rows where rightmost 8 chars are blank
}

OUT_HEADERS = [
    "STT", "Tên DN XNK", "Mã DN XNK", "Số tờ khai",
    "Tờ khai XK", "Trị giá tính thuế (USD)", "Thuế NK", "Ghi chú",
]

TNR = "Times New Roman"

# Region output files live beside this script's SKILL.md in regions.txt.
DEFAULT_REGIONS_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "regions.txt")
INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\\\]")


def deaccent(s):
    """Lowercase + strip Vietnamese accents (đ/Đ -> d) for matching."""
    if s is None:
        return ""
    s = str(s).replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower()


def load_regions(path):
    """Read sheet<TAB>term | term lines from a UTF-8 region config file."""
    regions = []
    seen = set()
    with open(path, encoding="utf-8") as f:
        for line_no, raw in enumerate(f, start=1):
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "\t" not in line:
                raise ValueError(f"{path}:{line_no}: expected sheet_name<TAB>keywords")
            stem, raw_terms = (part.strip() for part in line.split("\t", 1))
            if not stem or len(stem) > 31 or INVALID_SHEET_CHARS.search(stem):
                raise ValueError(f"{path}:{line_no}: invalid Excel sheet name {stem!r}")
            if stem.casefold() in seen:
                raise ValueError(f"{path}:{line_no}: duplicate sheet name {stem!r}")
            terms = [deaccent(term).strip() for term in raw_terms.split("|") if term.strip()]
            if not terms:
                raise ValueError(f"{path}:{line_no}: region needs at least one keyword")
            seen.add(stem.casefold())
            regions.append((stem, terms))
    if not regions:
        raise ValueError(f"{path}: no regions configured")
    return regions


def strip8(v):
    """Drop the first 8 characters of a value (returns '' if shorter)."""
    if v is None:
        return ""
    s = str(v)
    return s[8:] if len(s) > 8 else ""


def right8(v):
    """Return the rightmost 8 characters after trimming whitespace."""
    if v is None:
        return ""
    return str(v).strip()[-8:]


def to_float(v, stats=None):
    if v is None:
        return 0.0
    if isinstance(v, bool):
        if stats is not None:
            stats["unparseable_amounts"] = stats.get("unparseable_amounts", 0) + 1
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    if "." in s and "," in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", "")
    if not re.fullmatch(r"[+-]?(\d+(\.\d*)?|\.\d+)([eE][+-]?\d+)?", s):
        if stats is not None:
            stats["unparseable_amounts"] = stats.get("unparseable_amounts", 0) + 1
        return 0.0
    return float(s)


def cell_to_text(cell):
    """Return a cell value as text, preserving zero-padded numeric codes when possible."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)) and float(v).is_integer():
        s = str(int(v))
        fmt = str(cell.number_format or "")
        # Excel exports sometimes store Ma_DN_XNK as a number with a zero-only
        # display format (e.g. 0000000000). Preserve those leading zeroes.
        if fmt and set(fmt) == {"0"} and len(fmt) > len(s):
            return s.zfill(len(fmt))
        return s
    return str(v).strip()


def match_terms(addr, terms):
    """True if the (de-accented) address contains any configured term."""
    if not addr:
        return False
    s = deaccent(addr)
    return any(t in s for t in terms)


def choose_region(addr, regions):
    """Choose the region whose locality term appears furthest right in an address."""
    if not addr:
        return None
    s = deaccent(addr)
    candidates = []
    for stem, terms in regions:
        pos = max((s.rfind(term) for term in terms), default=-1)
        if pos >= 0:
            candidates.append((pos, stem))
    if not candidates:
        return None
    latest = max(pos for pos, _ in candidates)
    winners = [stem for pos, stem in candidates if pos == latest]
    return winners[0] if len(winners) == 1 else None


def split_regions(rows, regions):
    """Assign each row once; locality at the end of the address takes priority."""
    grouped = {stem: [] for stem, _ in regions}
    unmatched = []
    for row in rows:
        stem = choose_region(row["A"], regions)
        if stem is None:
            unmatched.append(row)
        else:
            grouped[stem].append(row)
    return grouped, unmatched


def norm_header(v):
    """Normalize header text for robust exact matching."""
    return deaccent(v).strip().replace(" ", "_")


def detect_src(ws, header_rows):
    """Use header names when present; fall back to legacy fixed positions."""
    detected = SRC.copy()
    header_row = max(1, header_rows)
    headers = {norm_header(ws.cell(header_row, c).value): c
               for c in range(1, ws.max_column + 1)
               if ws.cell(header_row, c).value is not None}
    wanted = {
        "So_to_khai": "so_to_khai",
        "Ma_LH": "ma_lh",
        "Ma_DN_XNK": "ma_dn_xnk",
        "Ten_DN_XNK": "ten_dn_xnk",
        "Ma_dia_chi": "ma_dia_chi_dn_xnk",
        "So_quan_ly": "so_quan_ly_cua_noi_bo_doanh_nghiep",
        "Tong_tri_gia": "tong_tri_gia_tinh_thue",
    }
    for key, header in wanted.items():
        if header in headers:
            detected[key] = headers[header]
    return detected


def step1(ws, ma_lh, header_rows, stats=None):
    """Filter by Ma_LH (one or more codes), remap columns, sort A->Z by name."""
    src = detect_src(ws, header_rows)
    rows = []
    if isinstance(ma_lh, str):
        targets = {x.strip().upper() for x in ma_lh.split(",") if x.strip()}
    else:
        targets = {str(x).strip().upper() for x in ma_lh}
    for r in range(header_rows + 1, ws.max_row + 1):
        mlh = ws.cell(r, src["Ma_LH"]).value
        if mlh is None:
            continue
        if str(mlh).strip().upper() not in targets:
            continue
        so_quan_ly = ws.cell(r, src["So_quan_ly"]).value
        to_khai_xk = strip8(so_quan_ly)
        if to_khai_xk == "":
            continue
        rows.append({
            "A": ws.cell(r, src["Ma_dia_chi"]).value,
            "B": ws.cell(r, src["Ten_DN_XNK"]).value,
            "C": cell_to_text(ws.cell(r, src["Ma_DN_XNK"])),
            "D": ws.cell(r, src["So_to_khai"]).value,
            "E": to_khai_xk,
            "F": to_float(ws.cell(r, src["Tong_tri_gia"]).value, stats),
        })
    rows.sort(key=lambda x: (str(x["B"]) if x["B"] is not None else "").lower())
    return rows


def write_step1(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Step1"
    ws.append(["A_Ma_dia_chi", "B_Ten_DN", "C_Ma_DN",
               "D_So_to_khai", "E_So_quan_ly", "F_Tri_gia"])
    for rec in rows:
        ws.append([rec["A"], rec["B"], rec["C"], rec["D"], rec["E"], rec["F"]])
    wb.save(path)


def bucket_rows_by_code(rows):
    """Bucket sorted rows by trimmed company code; blank codes stay separate."""
    buckets = []
    by_code = {}
    for rec in rows:
        code = str(rec["C"]).strip() if rec["C"] is not None else ""
        if code:
            group = by_code.get(code)
            if group is None:
                group = []
                by_code[code] = group
                buckets.append((code, group))
        else:
            group = []
            buckets.append(("", group))
        group.append(rec)
    return buckets


def build_region_sheet(ws, rows, terms, sheet_title, opts):
    hp = list(rows) if terms is None else [r for r in rows if match_terms(r["A"], terms)]
    # Sort for display, then bucket by code without relying on adjacency.
    hp.sort(key=lambda x: ((str(x["B"]) if x["B"] is not None else "").lower(),
                           str(x["C"]) if x["C"] is not None else ""))
    groups = bucket_rows_by_code(hp)
    ws.title = sheet_title

    # ---- styles ----
    font_title = Font(name=TNR, size=14, bold=True)
    font_hdr = Font(name=TNR, size=11, bold=True)
    font_data = Font(name=TNR, size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    a_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_hdr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    a_center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    a_right = Alignment(horizontal="right", vertical="center")

    mm = opts["month"] or "__"
    yyyy = opts["year"] or "____"
    title = (f"DANH SÁCH DOANH NGHIỆP LÀM THỦ TỤC NHẬP KHẨU "
             f"TẠI CHỖ THÁNG {mm}/{yyyy}")
    sub = (f"(Kèm theo Thông báo số {opts['tb_no'] or ''}/TB-XNKTC(GC) "
           f"ngày {opts['tb_day'] or '__'} tháng {opts['tb_month'] or '__'} "
           f"năm {opts['tb_year'] or '____'})")

    # ---- header rows 1-4 ----
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A1"].alignment = a_center
    ws["A2"] = sub
    ws["A2"].font = font_title
    ws["A2"].alignment = a_center
    # row 3 left blank
    for col, name in enumerate(OUT_HEADERS, start=1):
        c = ws.cell(4, col, name)
        c.font = font_hdr
        c.alignment = a_hdr
        c.border = border

    # ---- data rows ----
    r = 5
    stt = 0
    for code, group in groups:
        stt += 1
        grp_start = r
        for gi, rec in enumerate(group):
            if gi == 0:
                ws.cell(r, 1, stt)
                ws.cell(r, 2, rec["B"])
                ws.cell(r, 3, code)
            ws.cell(r, 4, rec["D"])
            ws.cell(r, 5, rec["E"])
            ws.cell(r, 6, rec["F"])
            ws.cell(r, 7, 0)
            ws.cell(r, 8, sheet_title if r == 5 else "")
            for col in range(1, 9):
                cell = ws.cell(r, col)
                cell.font = font_data
                cell.border = border
            ws.cell(r, 1).alignment = a_center
            ws.cell(r, 2).alignment = a_left_top
            ws.cell(r, 3).alignment = a_center_top
            ws.cell(r, 4).alignment = a_center
            ws.cell(r, 4).number_format = "0"
            ws.cell(r, 5).alignment = a_center
            ws.cell(r, 6).alignment = a_right
            ws.cell(r, 6).number_format = "#,##0.00"
            ws.cell(r, 7).alignment = a_center
            ws.cell(r, 8).alignment = a_center
            r += 1
        grp_end = r - 1
        if grp_end > grp_start:
            for col in (1, 2, 3):
                ws.merge_cells(start_row=grp_start, start_column=col,
                               end_row=grp_end, end_column=col)

    # ---- column widths ----
    widths = {"A": 6, "B": 34, "C": 16, "D": 18,
              "E": 18, "F": 18, "G": 12, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return len(hp), stt


def build_region_file(rows, out_path, terms, sheet_title, opts):
    wb = openpyxl.Workbook()
    n_rows, n_grp = build_region_sheet(wb.active, rows, terms, sheet_title, opts)
    wb.save(out_path)
    return n_rows, n_grp


def build_unmatched_sheet(ws, rows, opts):
    ws.title = "unmatched"
    headers = OUT_HEADERS + ["Địa chỉ nguồn", "Lý do"]
    font_title = Font(name=TNR, size=14, bold=True)
    font_hdr = Font(name=TNR, size=11, bold=True)
    font_data = Font(name=TNR, size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    a_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_hdr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    a_right = Alignment(horizontal="right", vertical="center")

    mm = opts["month"] or "__"
    yyyy = opts["year"] or "____"
    ws.merge_cells("A1:J1")
    ws["A1"] = f"DÒNG CHƯA KHỚP 9 NHÓM SAU KHI LỌC E21/G13 THÁNG {mm}/{yyyy}"
    ws["A1"].font = font_title
    ws["A1"].alignment = a_center
    for col, name in enumerate(headers, start=1):
        c = ws.cell(4, col, name)
        c.font = font_hdr
        c.alignment = a_hdr
        c.border = border

    data = sorted(rows, key=lambda x: ((str(x["B"]) if x["B"] is not None else "").lower(),
                                      str(x["C"]) if x["C"] is not None else ""))
    groups = bucket_rows_by_code(data)
    stt = 0
    r = 5
    for code, group in groups:
        stt += 1
        grp_start = r
        for gi, rec in enumerate(group):
            if gi == 0:
                ws.cell(r, 1, stt)
                ws.cell(r, 2, rec["B"])
                ws.cell(r, 3, code)
            ws.cell(r, 4, rec["D"])
            ws.cell(r, 5, rec["E"])
            ws.cell(r, 6, rec["F"])
            ws.cell(r, 7, 0)
            ws.cell(r, 8, "unmatched" if r == 5 else "")
            ws.cell(r, 9, rec["A"])
            ws.cell(r, 10, "no region match")
            for col in range(1, 11):
                cell = ws.cell(r, col)
                cell.font = font_data
                cell.border = border
            ws.cell(r, 1).alignment = a_center
            ws.cell(r, 2).alignment = a_left_top
            ws.cell(r, 3).alignment = a_center
            ws.cell(r, 4).alignment = a_center
            ws.cell(r, 4).number_format = "0"
            ws.cell(r, 5).alignment = a_center
            ws.cell(r, 6).alignment = a_right
            ws.cell(r, 6).number_format = "#,##0.00"
            ws.cell(r, 7).alignment = a_center
            ws.cell(r, 8).alignment = a_center
            ws.cell(r, 9).alignment = a_left_top
            ws.cell(r, 10).alignment = a_center
            r += 1
        grp_end = r - 1
        if grp_end > grp_start:
            for col in (1, 2, 3):
                ws.merge_cells(start_row=grp_start, start_column=col,
                               end_row=grp_end, end_column=col)

    widths = {"A": 6, "B": 34, "C": 16, "D": 18, "E": 18, "F": 18,
              "G": 12, "H": 14, "I": 60, "J": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    return len(data), stt


def write_summary_sheet(ws, step1_rows, counts):
    ws.title = "summary"
    headers = ["Sheet", "Terms", "Rows", "Companies", "Note"]
    font_hdr = Font(name=TNR, size=11, bold=True)
    font_data = Font(name=TNR, size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(["NKTC summary"])
    ws.append(["Step1 rows", step1_rows])
    ws.append(["Region rows", sum(n for stem, terms, n, _ in counts if terms)])
    ws.append(["Unmatched rows", sum(n for stem, terms, n, _ in counts if not terms)])
    ws.append(["Coverage", f"{(sum(n for stem, terms, n, _ in counts if terms) / step1_rows * 100) if step1_rows else 0:.2f}%"])
    ws.append([])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(7, c)
        cell.font = font_hdr
        cell.border = border
    for stem, terms, n_rows, n_grp in counts:
        ws.append([stem, ", ".join(terms) if terms else "", n_rows, n_grp, "not matched by 9 regions" if not terms else ""])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5):
        for cell in row:
            cell.font = font_data if cell.row != 7 else font_hdr
            if cell.row >= 7:
                cell.border = border
    widths = {"A": 16, "B": 48, "C": 12, "D": 12, "E": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_region_workbook(rows, out_path, opts, regions):
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    grouped, unmatched = split_regions(rows, regions)
    counts = []
    for stem, terms in regions:
        ws = wb.create_sheet()
        n_rows, n_grp = build_region_sheet(ws, grouped[stem], None, stem, opts)
        counts.append((stem, terms, n_rows, n_grp))
    ws = wb.create_sheet("unmatched")
    n_rows, n_grp = build_unmatched_sheet(ws, unmatched, opts)
    counts.append(("unmatched", [], n_rows, n_grp))
    write_summary_sheet(summary_ws, len(rows), counts)
    wb.save(out_path)
    return counts


def main():
    ap = argparse.ArgumentParser(description="NKTC Excel processor")
    ap.add_argument("input", help="source .xlsx")
    ap.add_argument("-o", "--output", default=None,
                    help="output .xlsx file for 9 sheets (default: <input-stem>_nktc.xlsx). "
                         "With --separate-files, this is the output directory.")
    ap.add_argument("--separate-files", action="store_true",
                    help="legacy mode: write 9 separate .xlsx files into output directory")
    ap.add_argument("--sheet", default=None, help="source sheet name")
    ap.add_argument("--ma-lh", default="E21,G13",
                    help="comma-separated Ma_LH codes to keep")
    ap.add_argument("--header-rows", type=int, default=1)
    ap.add_argument("--regions", default=DEFAULT_REGIONS_PATH,
                    help="UTF-8 config: sheet_name<TAB>keyword | keyword (default: regions.txt)")
    ap.add_argument("--step1-out", default=None,
                    help="optional path to also save the Step-1 table")
    ap.add_argument("--month", default=None)
    ap.add_argument("--year", default=None)
    ap.add_argument("--tb-no", default=None)
    ap.add_argument("--tb-day", default=None)
    ap.add_argument("--tb-month", default=None)
    ap.add_argument("--tb-year", default=None)
    args = ap.parse_args()
    try:
        regions = load_regions(args.regions)
    except (OSError, ValueError) as exc:
        ap.error(str(exc))

    # ---- auto-fill dates from today's date when not given ----
    # Title (row 1)    -> previous month + CURRENT year (the reported period)
    # Subtitle (row 2) -> current month + current year  (the issue date)
    today = datetime.date.today()
    first_of_month = today.replace(day=1)
    prev = first_of_month - datetime.timedelta(days=1)  # last day of prev month
    if args.month is None:
        args.month = f"{prev.month:02d}"
    if args.year is None:
        args.year = str(today.year)
    if args.tb_month is None:
        args.tb_month = f"{today.month:02d}"
    if args.tb_year is None:
        args.tb_year = str(today.year)

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    stats = {}
    rows = step1(ws, args.ma_lh, args.header_rows, stats)
    print(f"Step 1: {len(rows)} rows with Ma_LH in {args.ma_lh}")
    unparseable_amounts = stats.get("unparseable_amounts", 0)
    if unparseable_amounts:
        print(
            f"CẢNH BÁO: {unparseable_amounts} dòng có Trị giá không đọc được, "
            "đã ghi 0.00 - kiểm tra lại file nguồn."
        )

    if args.step1_out:
        write_step1(rows, args.step1_out)
        print(f"Step 1 table -> {args.step1_out}")

    opts = {k: getattr(args, k) for k in
            ("month", "year", "tb_no", "tb_day", "tb_month", "tb_year")}

    if args.separate_files:
        outdir = args.output or "."
        if outdir and not os.path.isdir(outdir):
            os.makedirs(outdir, exist_ok=True)

        print("Step 2: building per-region files")
        grouped, unmatched = split_regions(rows, regions)
        grand = 0
        for stem, terms in regions:
            out_path = os.path.join(outdir, f"{stem}.xlsx")
            n_rows, n_grp = build_region_file(grouped[stem], out_path, None, stem, opts)
            grand += n_rows
            print(f"  {stem}.xlsx: {n_rows} rows in {n_grp} companies "
                  f"(terms: {', '.join(terms)})")
        out_path = os.path.join(outdir, "unmatched.xlsx")
        wb_unmatched = openpyxl.Workbook()
        n_rows, n_grp = build_unmatched_sheet(wb_unmatched.active, unmatched, opts)
        wb_unmatched.save(out_path)
        grand += n_rows
        print(f"  unmatched.xlsx: {n_rows} rows in {n_grp} companies (not matched by configured regions)")
        region_rows = grand - len(unmatched)
        coverage = (region_rows / len(rows) * 100) if rows else 0
        print(f"Summary: step1={len(rows)} region_rows={region_rows} unmatched={len(unmatched)} coverage={coverage:.2f}%")
        print(f"Step 2 done: {grand} rows across {len(regions) + 1} files")
    else:
        out_path = args.output
        if out_path is None:
            stem = os.path.splitext(os.path.basename(args.input))[0]
            out_path = os.path.join(os.getcwd(), f"{stem}_nktc.xlsx")
        parent = os.path.dirname(os.path.abspath(out_path))
        if parent and not os.path.isdir(parent):
            os.makedirs(parent, exist_ok=True)

        print(f"Step 2: building summary + {len(regions)} region sheets + unmatched")
        counts = build_region_workbook(rows, out_path, opts, regions)
        grand = 0
        for stem, terms, n_rows, n_grp in counts:
            grand += n_rows
            if terms:
                print(f"  sheet {stem}: {n_rows} rows in {n_grp} companies "
                      f"(terms: {', '.join(terms)})")
            else:
                print(f"  sheet {stem}: {n_rows} rows in {n_grp} companies (not matched by 9 regions)")
        unmatched_rows = sum(n for _, terms, n, _ in counts if not terms)
        region_rows = grand - unmatched_rows
        coverage = (region_rows / len(rows) * 100) if rows else 0
        print(f"Summary: step1={len(rows)} region_rows={region_rows} unmatched={unmatched_rows} coverage={coverage:.2f}%")
        print(f"Step 2 done: {grand} rows across {len(regions) + 2} sheets -> {out_path}")


if __name__ == "__main__":
    sys.exit(main())
