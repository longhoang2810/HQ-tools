#!/usr/bin/env python3
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "nktc_process.py"
REGIONS = ROOT / "regions.txt"
REGIONS_34_BACKUP = ROOT / "regions-34-backup.txt"
sys.path.insert(0, str(ROOT / "scripts"))
from nktc_process import choose_region, load_regions


class RegionConfigTest(unittest.TestCase):
    def test_default_config_has_nine_regions_and_backup_has_34(self):
        regions = load_regions(REGIONS)
        backup = load_regions(REGIONS_34_BACKUP)
        self.assertEqual(len(regions), 34)
        self.assertIn("HCM", [name for name, _ in regions])
        self.assertEqual(len(backup), 34)
        self.assertIn("HCM", [name for name, _ in backup])

    def test_default_config_has_hcm_and_creates_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            output = Path(tmp) / "out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append([
                "So_to_khai", "Ma_LH", "Ma_DN_XNK", "Ten_DN_XNK",
                "Ma_dia_chi_DN_XNK", "So_quan_ly_cua_noi_bo_doanh_nghiep",
                "Tong_tri_gia_tinh_thue",
            ])
            ws.append([123456789, "E21", "0312345678", "DN HCM", "TP.HCM", "12345678XK01", 100])
            wb.save(source)
            subprocess.run(
                [sys.executable, str(SCRIPT), str(source), "-o", str(output), "--regions", str(REGIONS_34_BACKUP)],
                check=True, capture_output=True, text=True,
            )
            result = load_workbook(output, data_only=True)
            self.assertIn("HCM", result.sheetnames)
            self.assertEqual(result["HCM"]["B5"].value, "DN HCM")

    def test_last_locality_wins_over_street_name(self):
        regions = load_regions(REGIONS)
        self.assertEqual(
            choose_region("Đường Điện Biên Phủ, Thành phố Hải Phòng", regions),
            "hp",
        )
        self.assertEqual(
            choose_region("Số 437 Đường Đà Nẵng, Thành phố Hải Phòng", regions),
            "hp",
        )
        self.assertEqual(
            choose_region("Ngõ 43 đường Than Khê, tỉnh Bắc Ninh", regions),
            "BN",
        )


if __name__ == "__main__":
    unittest.main()
